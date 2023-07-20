import os
from openpyxl import load_workbook
import win32api

# Main folder path
main_folder_path = "C:/Path/To/Your/Main/Folder"

# Recursive function to process files in subfolders
def process_files(folder_path):
    # Get a list of all files and folders in the current folder
    entries = os.listdir(folder_path)
    
    # Iterate through each entry
    for entry in entries:
        # Create the full path of the current entry
        entry_path = os.path.join(folder_path, entry)
        
        # Check if the entry is a file
        if os.path.isfile(entry_path):
            # Check if the file is an Excel file
            if entry.endswith(".xlsx") or entry.endswith(".xls"):
                # Load the workbook
                workbook = load_workbook(entry_path)
                
                # Get a list of all sheets in the workbook
                sheets = workbook.sheetnames
                
                # Iterate through each sheet and print
                for sheet_name in sheets:
                    # Print the sheet using the default printer
                    win32api.ShellExecute(0, "print", entry_path, None, ".", 0)
                    
        # Check if the entry is a folder
        elif os.path.isdir(entry_path):
            # Recursively process files in the subfolder
            process_files(entry_path)

# Start processing files in the main folder
process_files(main_folder_path)
